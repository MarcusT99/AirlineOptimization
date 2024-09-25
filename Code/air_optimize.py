# -*- coding: utf-8 -*-
"""
03MAR24
Comp 2 Final Project
Airline price optimization
@author: David allen
"""

##### import statements ######################################################


# reading from excel
import os
import openpyxl
from openpyxl.utils.cell import range_boundaries

# searching for patterns
import re

# optimization specific (acknowledge * is not the pref technique)
from pyomo.environ import *

# may need to comment out, this is a non-standard optimization program
import gurobipy 


##### Index Notation #########################################################


J = set()       # set of all airports

I = set()       # set of all carriers, commonly called airlines

F = list()      # list of trips
'Fji'               ## trips have distinct names = "{sheet_name} {trip}"

T = dict()      # nest dict of airport (outer key): carrier (inner key)
'T[j,i]'            ## with list of all available trips (value)
                    ## intermediate notation used to build subsets

G = dict()      # subset of carriers by airports in T, 
'G ⊂ T'             ## (j,i) in T

H = dict()      # subset of trips by carrier by airport in T
'H ⊂ GxT'       

M = dict()      # dict of airport (key): dist in miles (value)
'M[j]'

D = dict()      # dict of carrier (key): cost of delay (value)
'D[i]'              ## for any leg
                    
B = dict()      # dict of carrier (key): cost of lost bag (value)
'B[i]'

P = dict()      # dict of trip (key): price (value)
'P[f]'

L = dict()      # dict of trip (key): number of legs (value)
'L[f]'

r = 0.67        # cost per mile traveled to any airport
# got from source below, used TDY travel rate for automobiles
# https://www.travel.dod.mil/Travel-Transportation-Rates/Mileage-Rates/

'X[j,i,f]'      # decision variable, binary, 1 if buy ticket
                    ## by airport, by carrier, by trip


##### Create Indexes from Excel File #########################################


### J

# path to excel file
path = 'air_LP_data_final.xlsx'

# load the workbook
wb = openpyxl.load_workbook(path)

### set of airports
# regular expression to match sheet names with exactly three capital letters
# sheets for data by departure airport have three letter IATA code as name
port_sheet = re.compile(r'^[A-Z]{3}$')

# iterate through the sheet names in the workbook
for sheet_name in wb.sheetnames:
    # if the sheet name matches the pattern, add it to the set
    if port_sheet.match(sheet_name):
        J.add(sheet_name)

#print(J)

### M

# access the sheet named 'port_info'
sheet = wb['port_info']

# find the column numbers for 'iata' and 'travel_miles'
iata_col = ''
travel_miles_col = ''
for col in range(1, sheet.max_column + 1):
    header_value = sheet.cell(row = 1, column=col).value
    if header_value == 'iata':
        iata_col = col
    elif header_value == 'travel_miles':
        travel_miles_col = col

    # stop after finding both columns
    if iata_col and travel_miles_col:
        break

for row in range(2, sheet.max_row + 1):
    iata_code = sheet.cell(row = row, column = iata_col).value
    travel_mile = sheet.cell(row = row, column = travel_miles_col).value
    # ensure float dtype and make k:v pairs    
    travel_mile_float = float(travel_mile)
    M[iata_code] = travel_mile_float

#print(M)

### I

# access the sheet named 'delay'
sheet = wb['delay']

# assumes the first row contains headers and 'carrier_name'
# valid assumption because output is specific to api and project
# find the column number for 'carrier_name'
carrier_name_col = ''
for col in range(1, sheet.max_column + 1):
    if sheet.cell(row = 1, column = col).value == 'carrier_name':
        carrier_name_col = col
        break
# iterate over the rows, starting from row 2 to skip the header
for row in range(2, sheet.max_row + 1):
    carrier_name = sheet.cell(row = row, column = carrier_name_col).value
    I.add(carrier_name)

#print(I)

### D

# find the column numbers for 'carrier_name' and 'expected_cost'
carrier_name_col = ''
expected_cost_col = ''
for col in range(1, sheet.max_column + 1):
    header_value = sheet.cell(row = 1, column = col).value
    if header_value == 'carrier_name':
        carrier_name_col = col
    elif header_value == 'expected_cost':
        expected_cost_col = col
    
    # stop after finding both cols
    if carrier_name_col and expected_cost_col:
        break

# iterate over rows
for row in range(2, sheet.max_row + 1):
    carrier_name = sheet.cell(row = row, column = carrier_name_col).value
    expected_cost = sheet.cell(row = row, column = expected_cost_col).value

    # attempt to convert the expected cost value to float
    expected_cost_float = float(expected_cost)
    D[carrier_name] = expected_cost_float
          
#print(D)

### B

sheet = wb['baggage']

# find the column numbers for 'carrier' and 'expected_cost'
carrier_col = ''
expected_cost_col = ''
for col in range(1, sheet.max_column + 1):
    header_value = sheet.cell(row = 1, column = col).value
    if header_value == 'CARRIER':
        carrier_col = col
    elif header_value == 'expected_cost':
        expected_cost_col = col

    # stop after finding both cols
    if carrier_col and expected_cost_col:
        break

for row in range(2, sheet.max_row + 1):
    carrier = sheet.cell(row = row, column = carrier_col).value
    expected_cost = sheet.cell(row = row, column = expected_cost_col).value
    expected_cost_float = float(expected_cost)
    B[carrier] = expected_cost_float

#print(B)  

### help from ChatGPT on building the indexes below
# prompt: 
# Write code to create a dictionary from an excel workbook. The dictionary 
# keys will be a concatenation of the sheet name and unique values of a column 
# labeled "airline". The dictionary values should be the values of the column 
# "Trip" concatenated with the sheet name, that have matching airline values 
# in their rows. The trip column has cells that are merged, because it was the 
# index in a multidimensional dataframe. 

### T_temp

# temporary notation as a stepping stone
T_temp = dict()

# Function to get value from possibly merged cell
def get_merged_cell_value1(sheet, row, col):
    cell_value = sheet.cell(row = row, column = col).value
    if cell_value is not None:
        return cell_value
    for range_ in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(range_))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return sheet.cell(row = min_row, column = min_col).value
    return None  # In case the cell is not merged or empty

# Iterate through each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Find the column numbers for 'airline' and 'Trip'
    airline_col = trip_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'airline':
            airline_col = col
        elif sheet.cell(row=1, column=col).value == 'Trip':
            trip_col = col
    
    # Check if both columns were found
    if airline_col and trip_col:
        # Iterate over the rows, starting from row 2 to skip the header
        for row in range(2, sheet.max_row + 1):
            airline = get_merged_cell_value1(sheet, row, airline_col)
            trip = get_merged_cell_value1(sheet, row, trip_col)
            
            # Only proceed if both airline and trip are not None
            if airline is not None and trip is not None:
                key = f"{sheet_name} {airline}"
                value = f"{sheet_name} {trip}"
                
                # Add or append the trip information to the dictionary
                if key in T_temp:
                    # Use a set to avoid duplicate values
                    T_temp[key].add(value)  
                else:
                    T_temp[key] = {value}

### T & F
# made from T_temp
T = dict()
F = list()
for j in J:
    for i in I:
        for key,value in T_temp.items():
            if j+' '+i == key:
                F.extend(value)
                if j in T:
                    nest_dict = T[j]
                    nest_dict[i] = value
                else:
                    T[j] = {i:value}

#print(T)
#print(F)

### G

for j,i in T.items():
    G[j] = i.keys()

#print(G)

### H

for j,i in T.items():
    for key,value in i.items():    
        H[key] = value
    
#print(H)

### P

# Function to handle potentially merged cells for the "Trip" column
def get_value_from_merged_cell2(sheet, row, col):
    cell = sheet.cell(row, column=col)
    if cell.value is not None:
        return cell.value
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Get the value from the first cell in the merged range
            first_cell = wb[sheet.title][merged_range.start_cell]
            return first_cell.value
    return None

# Iterate through each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Determine the column indices for 'Trip' and 'Trip_price'
    trip_col_index = trip_price_col_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Trip':
            trip_col_index = col
        elif sheet.cell(row=1, column=col).value == 'Trip_Price':
            trip_price_col_index = col
    
    # Proceed only if both columns were found
    if trip_col_index is not None and trip_price_col_index is not None:
        for row in range(2, sheet.max_row + 1):
            trip_value = get_merged_cell_value1(sheet, row, trip_col_index)
            trip_price_value = sheet.cell(row=row, column=trip_price_col_index).value
            
            # Skip rows where the 'Trip' value is missing
            if trip_value is not None:
                dict_key = f"{sheet_name} {trip_value}"
                trip_price_value = {trip_price_value}
                if dict_key not in P:
                    P[dict_key] = int(trip_price_value.pop())

# print(P)

### L

# Iterate through each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Determine the column indices for 'Trip' and 'Leg'
    trip_col_index = leg_col_index = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Trip':
            trip_col_index = col
        elif sheet.cell(row=1, column=col).value == 'Leg':
            leg_col_index = col
            
    # Proceed only if both necessary columns were found
    if trip_col_index is not None and leg_col_index is not None:
        for row in range(2, sheet.max_row + 1):
            trip_value = get_merged_cell_value1(sheet, row, trip_col_index)
            leg_value = sheet.cell(row=row, column=leg_col_index).value
            
            # Skip rows where either 'Trip' or 'Leg' value is missing
            if trip_value is not None and leg_value is not None:
                # Construct the dictionary key as 'TripValueSheetName'
                dict_key = f"{sheet_name} {trip_value}"
                if dict_key not in L:
                    L[dict_key] = [leg_value]
                else:
                   L[dict_key].append(leg_value)

# turn values into int of number of legs
for key in L.keys():
    L[key] = len(L[key])

#print(L)

wb.close()


##### Model Formulation ######################################################


# initialize model
BestFlights = ConcreteModel()

# define decision variables
BestFlights.X = Var(J,I,F, within = Binary)

# only want to buy one ticket constraint
def one_ticket(model):
    return sum(model.X[j,i,f] for j in J \
                              for i in G[j] \
                              for f in H[i]) == 1 
        
BestFlights.only_one = Constraint(rule = one_ticket)   

# optional constraint
# won't drive past certain dist constraint
'''
def long_drive(model,j,i,f):
    expr = model.X[j,i,f]*D[i] > D[i]
    return expr

BestFlights.drive = Constraint(J,G,H,rule = long_drive)
'''

# define objective function
# minimize total cost including hidden costs
def obj_rule(model):
    expr = sum(model.X[j,i,f]*(P[f] + r*M[j] + L[f]*(B[i] + D[i]))  \
               for j in J for i in G[j] for f in H[i])
    return expr
        
BestFlights.tot_cost = Objective(rule = obj_rule, sense = minimize) 
 

##### Solve for Result #1 ####################################################

# need to comment out if don't have gurobi, and use cbc line
opt = SolverFactory("gurobi") 
#opt = SolverFactory("cbc")

results = opt.solve(BestFlights, tee = True)
results.write()
BestFlights.display()

# print solution
from pyomo.environ import value
obj_val = round(value(BestFlights.tot_cost.expr),2)
for i in BestFlights.X:
    if BestFlights.X[i].value == 1:
        print("\nThe best flight is:",i[2][4:],"from",i[0])
        print("The airline operating the flight is:",i[1])
        print(f"The advertised price is ${P[i[2]]}",
              f"but the real cost is ${obj_val}")
        
##### Run Model Again for 2nd Best Flight ####################################
        
def add_exclusion_constraint(BestFlights, solution, constraint_id):
    # Create a unique constraint name using the provided constraint_id
    constraint_name = f"exclusion_constraint_{constraint_id}"
    
    # Define the exclusion constraint expression
    # Only include variables that are initialized and set to 1
    expr = sum(BestFlights.X[i] \
               for i in solution \
               if BestFlights.X[i].value is not None \
               and BestFlights.X[i].value == 1) <= len(solution) - 1
    
    # Add the constraint to the model with the unique name
    BestFlights.add_component(constraint_name, Constraint(expr=expr))
    
first_solution = [i for i in BestFlights.X if BestFlights.X[i].value \
                  is not None and BestFlights.X[i].value == 1]
    
add_exclusion_constraint(BestFlights, first_solution,1)

result = opt.solve(BestFlights)

# print solution
obj_val = round(value(BestFlights.tot_cost.expr),2)
for i in BestFlights.X:
    if BestFlights.X[i].value == 1:
        print("\nThe best flight is:",i[2][4:],"from",i[0])
        print("The airline operating the flight is:",i[1])
        print(f"The advertised price is ${P[i[2]]}",
              f"but the real cost is ${obj_val}")


##### Run Model Again for 3rd Best Flight ####################################


second_solution = [i for i in BestFlights.X if BestFlights.X[i].value \
                          is not None and BestFlights.X[i].value == 1]

add_exclusion_constraint(BestFlights, second_solution,2)

        
result = opt.solve(BestFlights)

       
# print solution
obj_val = round(value(BestFlights.tot_cost.expr),2)
for i in BestFlights.X:
    if BestFlights.X[i].value == 1:
        print("\nThe best flight is:",i[2][4:],"from",i[0])
        print("The airline operating the flight is:",i[1])
        print(f"The advertised price is ${P[i[2]]}",
              f"but the real cost is ${obj_val}")